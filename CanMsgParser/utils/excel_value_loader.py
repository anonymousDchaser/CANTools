# utils/excel_value_loader.py
"""从 Excel 矩阵文件加载信号值描述

支持多种实际格式：
- .xlsx (OpenXML，使用 openpyxl)
- .xls (OLE2 复合文档，使用 xlrd 1.x 或其他引擎)
- HTML 表格保存为 .xls（中国车企工具常见）
- CSV 保存为 .xls

自动检测文件真实格式，按策略依次尝试多种读取方式。

本模块提供两类接口：
- load_signal_value_db(file_path) -> dict[信号名, SignalValueInfo]
  返回每个信号的「富信息」：枚举矩阵(raw_choices)、总线/物理最大最小值、
  精度/偏移/单位，并能对「无明确枚举、仅有数值范围」的信号生成边界测试用的
  mock 原始值（最小/最大/中点等）。
- load_value_descriptions(file_path) -> dict[信号名, dict[int, str]]
  兼容旧调用方（曲线/监控页），仅返回枚举矩阵部分。
"""

import os
import re
from dataclasses import dataclass, field

import pandas as pd
import xlrd


# 信号名称列的匹配关键词（按优先级排列）
SIGNAL_NAME_KEYWORDS = [
    "signal name",       # 英文
    "signal_name",
    "信号名",
    "信号名称",
    "signal",            # 更宽泛的匹配
]

# 信号值描述列的匹配关键词（按优先级排列）
VALUE_DESC_KEYWORDS = [
    "signal value description",
    "信号值描述",
    "value description",
    "值描述",
    "value desc",
    "信号描述",          # 备选
    "signal description",
]

# 物理最小值 / 最大值 列
PHYS_MIN_KEYWORDS = ["物理最小值", "signal min. value (phys)", "physical min", "最小值", "min value"]
PHYS_MAX_KEYWORDS = ["物理最大值", "signal max. value(phys)", "physical max", "最大值", "max value"]

# 总线（hex）最小值 / 最大值 列 —— 这是原始值的真实取值范围，范围 mock 优先用此
BUS_MIN_KEYWORDS = ["总线最小值", "signal min. value (hex)", "bus min", "min value(hex)"]
BUS_MAX_KEYWORDS = ["总线最大值", "signal max. value(hex)", "bus max", "max value(hex)"]

# 精度 / 偏移 / 单位
SCALE_KEYWORDS = ["resolution", "精度", "scale"]
OFFSET_KEYWORDS = ["offset", "偏移量"]
UNIT_KEYWORDS = ["unit", "单位"]


@dataclass
class SignalValueInfo:
    """单个信号在 Excel 矩阵中的「富信息」。

    raw_choices : 明确枚举矩阵 {原始值(int): 描述(str)}，如 {0:"OFF",1:"ON"}
    bus_min/bus_max : 总线（hex）最小值/最大值，即原始值的真实取值范围
    phys_min/phys_max : 物理最小值/最大值
    scale/offset/unit : 精度/偏移/单位（来自矩阵，用于把原始值换算成物理量做标签）
    desc_text : 原始信号值描述文本
    """

    raw_choices: dict = field(default_factory=dict)
    bus_min: int | None = None
    bus_max: int | None = None
    phys_min: float | None = None
    phys_max: float | None = None
    scale: float = 1.0
    offset: float = 0.0
    unit: str = ""
    desc_text: str = ""

    @property
    def has_enum(self) -> bool:
        return bool(self.raw_choices)

    def mock_choices(self) -> dict | None:
        """对「无明确枚举、仅有数值范围」的信号，生成边界测试用的原始值下拉项。

        边界测试规则：始终包含 最小 / 最大 / 中点，范围较大时再加 1/3、2/3 等分点；
        优先使用总线（hex）最小/最大值作为原始值范围，缺失时退回用物理范围经
        精度/偏移换算成原始值。返回 {原始值(int): 标签(str)}，无范围信息时返回 None。
        """
        lo, hi = self._raw_range()
        if lo is None or hi is None or hi < lo:
            return None
        span = hi - lo
        if span <= 0:
            return {lo: self._label(lo, "最小")}

        # 候选点：最小 / 最大 / 中点 必含；范围较大再补两个等分点
        cand: list[tuple[int, str]] = [
            (lo, "最小"),
            (hi, "最大"),
            ((lo + hi) // 2, "中间"),
        ]
        if span >= 4:
            cand.append((lo + span // 3, ""))
            cand.append((lo + 2 * span // 3, ""))

        seen: set[int] = set()
        ordered: list[int] = []
        for raw, _ in cand:
            if raw not in seen:
                seen.add(raw)
                ordered.append(raw)

        # 数量过多时均匀裁剪，但保留 最小/最大/中间
        if len(ordered) > 7:
            must = {lo, hi, (lo + hi) // 2}
            others = [r for r in sorted(seen) if r not in must]
            keep = set(must)
            if others:
                step = max(1, len(others) // (7 - len(must)))
                keep |= set(others[::step][: (7 - len(must))])
            ordered = sorted(keep)

        # 重新生成带角色标签的字典（中点/最小/最大带中文角色）
        role_of = {raw: role for raw, role in cand}
        return {raw: self._label(raw, role_of.get(raw, "")) for raw in ordered}

    def _raw_range(self) -> tuple[int | None, int | None]:
        lo, hi = self.bus_min, self.bus_max
        if lo is not None and hi is not None:
            return lo, hi
        # 退回：用物理最小值/最大值 + 精度/偏移 换算成原始值
        if (
            self.phys_min is not None
            and self.phys_max is not None
            and self.scale not in (None, 0)
        ):
            try:
                lo = int(round((self.phys_min - self.offset) / self.scale))
                hi = int(round((self.phys_max - self.offset) / self.scale))
                return lo, hi
            except Exception:
                return None, None
        return None, None

    def _label(self, raw: int, role: str) -> str:
        if self.scale not in (None, 0):
            phys = raw * self.scale + self.offset
            unit = f" {self.unit}" if self.unit else ""
            base = f"{phys:g}{unit}"
        else:
            base = f"原值{raw}"
        if role:
            return f"{role} {base}"
        return base


def load_signal_value_db(file_path: str) -> dict[str, SignalValueInfo]:
    """加载 Excel 矩阵文件，返回 {信号名: SignalValueInfo}。

    自动检测实际格式（.xlsx / .xls / HTML / CSV），遍历所有 sheet 找到含
    「信号名称 + 信号值描述」列的表。除枚举外，还会读取精度/偏移/单位、
    物理最小最大值、总线（hex）最小最大值，用于对无明确枚举的信号生成边界
    测试 mock 值。
    """
    ext = os.path.splitext(file_path)[1].lower()
    df = _try_read_excel(file_path, ext)
    if df is None:
        _raise_format_error(file_path)

    name_col = _find_column(df, SIGNAL_NAME_KEYWORDS)
    desc_col = _find_column(df, VALUE_DESC_KEYWORDS)
    if name_col is None:
        raise ValueError(
            f"未找到信号名称列。当前列名: {list(df.columns)}\n"
            f"尝试的关键词: {SIGNAL_NAME_KEYWORDS}"
        )
    if desc_col is None:
        raise ValueError(
            f"未找到信号值描述列。当前列名: {list(df.columns)}\n"
            f"尝试的关键词: {VALUE_DESC_KEYWORDS}"
        )

    phys_min_col = _find_column(df, PHYS_MIN_KEYWORDS)
    phys_max_col = _find_column(df, PHYS_MAX_KEYWORDS)
    bus_min_col = _find_column(df, BUS_MIN_KEYWORDS)
    bus_max_col = _find_column(df, BUS_MAX_KEYWORDS)
    scale_col = _find_column(df, SCALE_KEYWORDS)
    offset_col = _find_column(df, OFFSET_KEYWORDS)
    unit_col = _find_column(df, UNIT_KEYWORDS)

    db: dict[str, SignalValueInfo] = {}
    for _, row in df.iterrows():
        sig_name = str(row[name_col]).strip() if pd.notna(row[name_col]) else ""
        if not sig_name or sig_name == "nan":
            continue
        desc_text = str(row[desc_col]) if pd.notna(row[desc_col]) else ""
        parsed = parse_value_description(desc_text)

        info = db.get(sig_name, SignalValueInfo())
        if parsed:
            # 同一信号多行时合并枚举
            info.raw_choices.update(parsed)
        info.desc_text = desc_text

        if bus_min_col is not None:
            v = _to_int_raw(row[bus_min_col])
            if v is not None:
                info.bus_min = v
        if bus_max_col is not None:
            v = _to_int_raw(row[bus_max_col])
            if v is not None:
                info.bus_max = v
        if phys_min_col is not None:
            v = _to_float(row[phys_min_col])
            if v is not None:
                info.phys_min = v
        if phys_max_col is not None:
            v = _to_float(row[phys_max_col])
            if v is not None:
                info.phys_max = v
        if scale_col is not None:
            v = _to_float(row[scale_col])
            if v is not None:
                info.scale = v
        if offset_col is not None:
            v = _to_float(row[offset_col])
            if v is not None:
                info.offset = v
        if unit_col is not None:
            u = str(row[unit_col]).strip() if pd.notna(row[unit_col]) else ""
            if u and u != "nan":
                info.unit = u

        db[sig_name] = info

    return db


def load_value_descriptions(file_path: str) -> dict[str, dict[int, str]]:
    """兼容旧接口：仅返回 {信号名: {原始值: 描述}} 枚举部分。"""
    db = load_signal_value_db(file_path)
    return {name: dict(info.raw_choices) for name, info in db.items()}


# ════════════════════════ 数值解析辅助 ════════════════════════

def _to_int_raw(cell) -> int | None:
    """把单元格值解析为原始整数（支持 0x 十六进制、普通整数/小数）。"""
    if cell is None or (isinstance(cell, float) and pd.isna(cell)):
        return None
    s = str(cell).strip()
    if not s or s.lower() == "nan":
        return None
    try:
        if s.lower().startswith("0x"):
            return int(s, 16)
        return int(float(s))
    except (ValueError, TypeError):
        return None


def _to_float(cell) -> float | None:
    if cell is None or (isinstance(cell, float) and pd.isna(cell)):
        return None
    s = str(cell).strip()
    if not s or s.lower() == "nan":
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def parse_value_description(desc_text: str) -> dict[int, str]:
    """解析信号值描述文本为枚举矩阵 {原始值(int): 描述(str)}。

    支持多种格式（半角/全角冒号、等号、横杠、波浪号分隔）：
      - "0:OFF, 1:ON"
      - "0=OFF, 1=ON"
      - "0 - OFF, 1 - ON"
      - "0x0：Not Active, 0x1：ON"  （全角冒号）
      - "0-OFF; 1-ON"
      - 中英文逗号/分号/换行分隔

    注意：全角冒号「：」与全角波浪号「～」也纳入分隔符，以兼容中国车企矩阵。
    仅含范围（如 "0~255"）而无明确枚举的描述不会在此解析为枚举，交由范围 mock
    逻辑处理。
    """
    result: dict[int, str] = {}
    if not desc_text or pd.isna(desc_text):
        return result

    desc_text = str(desc_text).strip()
    if not desc_text:
        return result

    # 按逗号、分号、换行、全角逗号/分号等常见分隔符拆分
    pairs = re.split(r'[,;，；\n\r]+', desc_text)

    for pair in pairs:
        pair = pair.strip()
        if not pair:
            continue
        # 匹配 "数值 分隔符 描述" 模式
        # 分隔符支持：冒号(半/全角)、等号、各种横杠（-–—）、波浪号(半/全角 ～~)
        match = re.match(
            r'^(0[xX][\da-fA-F]+|\d+(?:\.\d+)?)\s*[:：=~\-–—~]\s*(.+)$',
            pair,
        )
        if match:
            val_str = match.group(1)
            desc = match.group(2).strip()
            try:
                if val_str.lower().startswith("0x"):
                    val = int(val_str, 16)
                else:
                    val = int(float(val_str))
                result[val] = desc
            except ValueError:
                pass

    return result


def parse_range_description(desc_text: str) -> tuple[float, float] | None:
    """从描述文本中解析「纯范围」(如 "0~255"、"0至100"、"0到100"、"0-100")。

    返回 (最小值, 最大值) 浮点；若文本不是纯范围则返回 None。
    仅用于描述本身写成了范围、且枚举解析为空时的兜底。
    """
    if not desc_text or pd.isna(desc_text):
        return None
    s = str(desc_text).strip()
    if not s:
        return None
    # 含明确的 "值:描述" 形式的不算纯范围
    if re.search(r'[:：=]', s):
        return None
    m = re.search(r'(-?\d+(?:\.\d+)?)\s*[~～至到\-–—]\s*(-?\d+(?:\.\d+)?)', s)
    if m:
        try:
            lo = float(m.group(1))
            hi = float(m.group(2))
            return (lo, hi)
        except ValueError:
            return None
    return None


# ════════════════════════ 文件读取（多格式自动探测） ════════════════════════

def _read_xls_via_xlrd_directly(file_path: str, header_row: int = 0) -> pd.DataFrame | None:
    """直接使用 xlrd 读取 OLE2 .xls 文件并转为 DataFrame。

    绕过 pandas 的 xlrd 版本限制（pandas>=3.x 要求 xlrd>=2.0.1，
    但 xlrd 2.x 已移除 .xls OLE2 支持），直接调用 xlrd 1.2.0 API。

    CAN矩阵文件通常有多个sheet（Macro1、Cover、History、Legend、Matrix），
    数据在 "Matrix" sheet中。因此需要遍历所有sheet查找包含目标列的那个。
    """
    try:
        wb = xlrd.open_workbook(file_path)
    except Exception:
        return None

    for ws in wb.sheets():
        if ws.nrows == 0:
            continue

        all_rows = []
        for row_idx in range(ws.nrows):
            row_data = []
            for col_idx in range(ws.ncols):
                cell = ws.cell(row_idx, col_idx)
                if cell.ctype == 0:
                    row_data.append('')
                elif cell.ctype == 2:
                    if cell.value == int(cell.value):
                        row_data.append(str(int(cell.value)))
                    else:
                        row_data.append(str(cell.value))
                elif cell.ctype == 3:
                    date_tuple = xlrd.xldate_as_tuple(cell.value, wb.datemode)
                    row_data.append(str(date_tuple))
                elif cell.ctype == 4:
                    row_data.append(str(bool(cell.value)))
                elif cell.ctype == 5:
                    row_data.append('')
                else:
                    row_data.append(str(cell.value))
            all_rows.append(row_data)

        for hr in range(min(5, len(all_rows))):
            headers = all_rows[hr]
            data_start = hr + 1
            data_rows = all_rows[data_start:]
            if not data_rows:
                continue
            max_cols = len(headers)
            for i in range(len(data_rows)):
                if len(data_rows[i]) < max_cols:
                    data_rows[i].extend([''] * (max_cols - len(data_rows[i])))
                elif len(data_rows[i]) > max_cols:
                    data_rows[i] = data_rows[i][:max_cols]
            df = pd.DataFrame(data_rows, columns=headers)
            if _has_target_columns(df):
                return df

    return None


def _try_read_excel(file_path: str, ext: str) -> pd.DataFrame | None:
    """尝试多种方式读取 Excel 文件，返回包含目标列的 DataFrame。"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            for header_row in [0, 1, 2, 3]:
                try:
                    df = pd.read_excel(
                        file_path, engine='openpyxl',
                        sheet_name=sheet_name, header=header_row
                    )
                    if _has_target_columns(df):
                        wb.close()
                        return df
                except Exception:
                    pass
        wb.close()
    except Exception:
        pass

    try:
        df = _read_xls_via_xlrd_directly(file_path)
        if df is not None and _has_target_columns(df):
            return df
    except Exception:
        pass

    for header_row in [0, 1]:
        try:
            df = pd.read_excel(file_path, engine='xlrd', header=header_row)
            if _has_target_columns(df):
                return df
        except Exception:
            pass

    for header_row in [0, 1]:
        try:
            df = pd.read_excel(file_path, header=header_row)
            if _has_target_columns(df):
                return df
        except Exception:
            pass

    for encoding in ['utf-8', 'gbk', 'gb2312', 'gb18030']:
        try:
            tables = pd.read_html(file_path, encoding=encoding)
            for df in tables:
                if _has_target_columns(df):
                    return df
        except Exception:
            pass

    for encoding in ['utf-8', 'gbk', 'gb2312', 'gb18030', 'utf-8-sig']:
        for sep in [',', '\t', ';']:
            try:
                df = pd.read_csv(file_path, encoding=encoding, sep=sep)
                if _has_target_columns(df):
                    return df
            except Exception:
                pass

    try:
        df_raw = pd.read_excel(file_path, header=None)
        if len(df_raw) > 0:
            for i in range(min(5, len(df_raw))):
                test_df = df_raw.copy()
                test_df.columns = test_df.iloc[i].astype(str)
                test_df = test_df.iloc[i + 1:].reset_index(drop=True)
                if _has_target_columns(test_df):
                    return test_df
    except Exception:
        pass

    for encoding in ['utf-8', 'gbk', 'gb18030']:
        try:
            tables = pd.read_html(file_path, encoding=encoding, header=None)
            for df_raw in tables:
                if len(df_raw) > 0:
                    for i in range(min(5, len(df_raw))):
                        test_df = df_raw.copy()
                        test_df.columns = test_df.iloc[i].astype(str)
                        test_df = test_df.iloc[i + 1:].reset_index(drop=True)
                        if _has_target_columns(test_df):
                            return test_df
        except Exception:
            pass

    return None


def _has_target_columns(df: pd.DataFrame) -> bool:
    name_col = _find_column(df, SIGNAL_NAME_KEYWORDS)
    desc_col = _find_column(df, VALUE_DESC_KEYWORDS)
    return name_col is not None and desc_col is not None


def _raise_format_error(file_path: str) -> None:
    try:
        with open(file_path, 'rb') as f:
            header_bytes = f.read(200)

        lower_bytes = header_bytes.lower()
        if b'<html' in lower_bytes or b'<table' in lower_bytes or b'<tr' in lower_bytes:
            actual_format = "HTML 表格（伪装为 .xls）"
        elif header_bytes[:2] == b'PK':
            actual_format = "ZIP/OpenXML (.xlsx)"
        elif header_bytes[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
            actual_format = "OLE2 复合文档 (.xls)"
            actual_format += " —— xlrd 直接读取失败，请检查文件是否损坏"
        elif header_bytes[:3] == b'\xef\xbb\xbf' or b',' in header_bytes[:100] or b'\t' in header_bytes[:100]:
            actual_format = "CSV/文本文件（伪装为 .xls）"
        else:
            actual_format = f"未知格式 (首字节: {header_bytes[:8].hex()})"

        header_preview = header_bytes[:80].decode('utf-8', errors='replace')

        raise ValueError(
            f"无法读取文件 {os.path.basename(file_path)}。\n"
            f"检测到的实际格式: {actual_format}\n"
            f"已尝试策略: openpyxl(xlsx), xlrd直接读取, pandas xlrd引擎, 默认引擎, HTML表格, CSV, 无表头推断\n"
            f"请确认文件未损坏且格式正确。\n\n"
            f"文件头部预览: {header_preview}..."
        )
    except ValueError:
        raise
    except Exception as e2:
        raise ValueError(f"无法读取文件: {e2}")


def _find_column(df: pd.DataFrame, keywords: list[str]) -> str | None:
    """在 DataFrame 列名中模糊查找。

    处理双语表头（换行分隔）、多余空格、大小写差异、半/全角差异。
    按关键词优先级返回第一个匹配的列名。
    """
    col_map = {}
    for col in df.columns:
        normalized = str(col).lower()
        normalized = re.sub(r'[\n\r]+', ' ', normalized)
        normalized = re.sub(r'\s+', ' ', normalized).strip()
        col_map[normalized] = col

    for kw in keywords:
        kw_lower = kw.lower().strip()
        if kw_lower in col_map:
            return col_map[kw_lower]
        for normalized, original in col_map.items():
            if kw_lower in normalized:
                return original

    return None
